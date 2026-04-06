"""数据查询工具 - 根据尺寸和重量查询名称"""

import sys
import os
from openpyxl import load_workbook
from PySide6.QtWidgets import (
    QApplication, QWidget, QVBoxLayout, QHBoxLayout,
    QLabel, QLineEdit, QPushButton, QTableWidget,
    QTableWidgetItem, QHeaderView, QMessageBox,
)
from PySide6.QtCore import Qt
from PySide6.QtGui import QFont


def parse_range(text):
    """解析范围字符串，返回 (min, max) 元组。
    单值如 "0.38" 返回 (0.38, 0.38)，
    范围如 "0.29-0.33" 返回 (0.29, 0.33)。
    解析失败返回 None。
    """
    if not text or not text.strip():
        return None
    text = text.strip()
    # 处理负数开头的范围（如 "1.886--2.999" 不太可能，但防御性处理）
    parts = text.split("-")
    nums = []
    i = 0
    while i < len(parts):
        part = parts[i].strip()
        if part == "" and i + 1 < len(parts):
            # 负号：合并下一段
            i += 1
            nums.append(-float(parts[i].strip()))
        else:
            nums.append(float(part))
        i += 1
    if len(nums) == 1:
        return (nums[0], nums[0])
    elif len(nums) >= 2:
        return (min(nums), max(nums))
    return None


def load_data(xlsx_path):
    """从 Excel 文件加载数据，返回记录列表"""
    wb = load_workbook(xlsx_path, read_only=True)
    ws = wb.active
    records = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or not row[1]:
            continue
        seq = str(row[0] or "").strip()
        name = str(row[1] or "").strip()
        size_text = str(row[2] or "").strip()
        weight_text = str(row[3] or "").strip()
        records.append({
            "seq": seq,
            "name": name,
            "size_text": size_text,
            "weight_text": weight_text,
            "size_range": parse_range(size_text),
            "weight_range": parse_range(weight_text),
        })
    wb.close()
    return records


def in_range(value, range_tuple):
    """判断值是否在范围内（含边界），范围为 None 表示该记录缺少此字段，不匹配"""
    if range_tuple is None:
        return False
    return range_tuple[0] <= value <= range_tuple[1]


class MainWindow(QWidget):
    """主窗口"""

    def __init__(self):
        super().__init__()
        self.setWindowTitle("数据查询工具")
        self.setMinimumSize(600, 500)

        # 加载数据
        xlsx_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data.xlsx")
        self.records = load_data(xlsx_path)

        self._init_ui()

    def _init_ui(self):
        """初始化界面"""
        layout = QVBoxLayout(self)

        # 输入区域
        input_layout = QHBoxLayout()
        font = QFont("微软雅黑", 10)

        label_size = QLabel("尺寸：")
        label_size.setFont(font)
        self.input_size = QLineEdit()
        self.input_size.setFont(font)
        self.input_size.setPlaceholderText("请输入尺寸，如 0.30")

        label_weight = QLabel("重量：")
        label_weight.setFont(font)
        self.input_weight = QLineEdit()
        self.input_weight.setFont(font)
        self.input_weight.setPlaceholderText("请输入重量，如 6.0")

        btn_query = QPushButton("查询")
        btn_query.setFont(font)
        btn_query.setFixedWidth(80)
        btn_query.clicked.connect(self._on_query)

        # 回车触发查询
        self.input_size.returnPressed.connect(self._on_query)
        self.input_weight.returnPressed.connect(self._on_query)

        input_layout.addWidget(label_size)
        input_layout.addWidget(self.input_size)
        input_layout.addWidget(label_weight)
        input_layout.addWidget(self.input_weight)
        input_layout.addWidget(btn_query)

        layout.addLayout(input_layout)

        # 结果表格
        self.table = QTableWidget()
        self.table.setColumnCount(4)
        self.table.setHorizontalHeaderLabels(["序号", "名称", "尺寸", "重量"])
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.table.setSelectionBehavior(QTableWidget.SelectRows)
        self.table.setFont(font)
        layout.addWidget(self.table)

        # 状态栏
        self.status_label = QLabel(f"共 {len(self.records)} 条数据，请输入条件后点击查询")
        self.status_label.setFont(font)
        layout.addWidget(self.status_label)

    def _on_query(self):
        """执行查询"""
        size_text = self.input_size.text().strip()
        weight_text = self.input_weight.text().strip()

        # 至少输入一个条件
        if not size_text and not weight_text:
            QMessageBox.warning(self, "提示", "请至少输入尺寸或重量中的一个条件")
            return

        # 解析输入值
        size_val = None
        weight_val = None
        try:
            if size_text:
                size_val = float(size_text)
        except ValueError:
            QMessageBox.warning(self, "提示", "尺寸格式不正确，请输入数字")
            return
        try:
            if weight_text:
                weight_val = float(weight_text)
        except ValueError:
            QMessageBox.warning(self, "提示", "重量格式不正确，请输入数字")
            return

        # 匹配
        results = []
        for rec in self.records:
            size_match = (size_val is None) or in_range(size_val, rec["size_range"])
            weight_match = (weight_val is None) or in_range(weight_val, rec["weight_range"])
            if size_match and weight_match:
                results.append(rec)

        # 更新表格
        self.table.setRowCount(len(results))
        for row, rec in enumerate(results):
            self.table.setItem(row, 0, QTableWidgetItem(rec["seq"]))
            self.table.setItem(row, 1, QTableWidgetItem(rec["name"]))
            self.table.setItem(row, 2, QTableWidgetItem(rec["size_text"]))
            self.table.setItem(row, 3, QTableWidgetItem(rec["weight_text"]))

        # 更新状态
        if results:
            self.status_label.setText(f"匹配到 {len(results)} 条结果")
        else:
            self.status_label.setText("未找到匹配的结果")


if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = MainWindow()
    window.show()
    sys.exit(app.exec())
